import streamlit as st
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Alignment, numbers
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64
from datetime import datetime

# 设置页面配置
st.set_page_config(
    page_title="Excel订单数据分类工具",
    page_icon="📊",
    layout="wide"
)

def extract_conductor_type(product_name):
    """根据产品名称判断导体类型。包含 "LV" 为铝导体，否则为铜导体。"""
    if "LV" in product_name:
        return "铝导体"
    else:
        return "铜导体"

def extract_cross_section(product_name):
    """从产品名称中提取横截面积。"""
    try:
        match = re.search(r'-\d+×(\d+)(?!\d)', product_name)
        if match:
            captured_group = match.group(1)
            return int(captured_group)
        return None
    except Exception as e:
        st.error(f"提取横截面积失败：{e}")
        return None

def create_or_load_workbook(filename):
    """创建或加载 Excel 工作簿。如果文件不存在，则创建新文件。"""
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    return workbook

def set_column_width(sheet):
    """自动调整列宽以适应内容。"""
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

def set_date_format(cell):
    """设置单元格的日期格式为 yyyy/mm/dd。"""
    cell.number_format = 'yyyy/mm/dd'

def write_data_to_excel(workbook, sheet_name, data, headers):
    """将数据写入 Excel 工作表，并处理日期格式。"""
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet()
        sheet.title = sheet_name
        sheet.append(headers)
        # 设置标题行样式
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in data:
        sheet.append(row)
        # 设置数据行样式
        for cell in sheet[sheet.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 处理日期格式
        for col_idx, header in enumerate(headers):
            if header in ["生产日期", "订单日期", "交期"]:
                cell = sheet.cell(row=sheet.max_row, column=col_idx + 1)
                if isinstance(cell.value, datetime):
                    set_date_format(cell)
    set_column_width(sheet)

def save_workbook_to_buffer(workbook):
    """将 Excel 工作簿保存到内存中的字节缓冲区。"""
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

def process_excel_data(uploaded_file):
    try:
        df = pd.read_excel(uploaded_file, sheet_name="总订单")
        df.dropna(how='all', inplace=True)

        total_rows = len(df)
        processed_count = 0
        errors = []
        
        # 创建新的工作簿
        aluminum_workbook = create_or_load_workbook("铝导体安排.xlsx")
        copper_workbook = create_or_load_workbook("铜导体安排.xlsx")

        # 删除默认的 Sheet 工作表
        if 'Sheet' in aluminum_workbook.sheetnames:
            del aluminum_workbook['Sheet']
        if 'Sheet' in copper_workbook.sheetnames:
            del copper_workbook['Sheet']

        headers = ["产品编码", "工单", "生产日期", "订单日期", "订单", "单位名称", 
                  "产品名称", "型号", "导体米数", "分排", "交期", "绝缘", "成缆", "外护", "交期"]
        
        # 创建错误处理工作表
        aluminum_error_sheet_name = "铝导体错误数据"
        copper_error_sheet_name = "铜导体错误数据"
        write_data_to_excel(aluminum_workbook, aluminum_error_sheet_name, [], headers + ["错误信息"])
        write_data_to_excel(copper_workbook, copper_error_sheet_name, [], headers + ["错误信息"])

        progress_bar = st.progress(0)
        status_text = st.empty()

        for index, row in df.iterrows():
            try:
                if pd.isna(row["产品名称"]):
                    errors.append(f"第 {index + 2} 行：产品名称为空，无法处理。")
                    # 将空产品名称的行写入错误工作表
                    data_row = [
                        row["产品编码"], row["工单"], row["生产日期"], row["订单日期"], 
                        row["订单"], row["单位名称"], "", row["型号"], 
                        row["数量"], row["分排"], row["交期"], row["绝缘"], 
                        row["成缆"], row["外护"], row["交期"], "产品名称为空"
                    ]
                    conductor_type = extract_conductor_type(str(row["产品名称"]) if not pd.isna(row["产品名称"]) else "")
                    if conductor_type == "铝导体":
                        write_data_to_excel(aluminum_workbook, aluminum_error_sheet_name, [data_row], headers + ["错误信息"])
                    elif conductor_type == "铜导体":
                        write_data_to_excel(copper_workbook, copper_error_sheet_name, [data_row], headers + ["错误信息"])
                    continue
                
                product_name = str(row["产品名称"])
                conductor_type = extract_conductor_type(product_name)
                cross_section = extract_cross_section(product_name)

                if cross_section is None:
                    errors.append(f"第 {index + 2} 行：无法提取横截面积，产品名称：{product_name}")
                    # 将无法提取横截面积的行写入错误工作表
                    data_row = [
                        row["产品编码"], row["工单"], row["生产日期"], row["订单日期"], 
                        row["订单"], row["单位名称"], product_name, row["型号"], 
                        row["数量"], row["分排"], row["交期"], row["绝缘"], 
                        row["成缆"], row["外护"], row["交期"], "无法提取横截面积"
                    ]
                    if conductor_type == "铝导体":
                        write_data_to_excel(aluminum_workbook, aluminum_error_sheet_name, [data_row], headers + ["错误信息"])
                    elif conductor_type == "铜导体":
                        write_data_to_excel(copper_workbook, copper_error_sheet_name, [data_row], headers + ["错误信息"])
                    continue

                data_row = [
                    row["产品编码"], row["工单"], row["生产日期"], row["订单日期"], 
                    row["订单"], row["单位名称"], product_name, row["型号"], 
                    row["数量"], row["分排"], row["交期"], row["绝缘"], 
                    row["成缆"], row["外护"], row["交期"]
                ]

                # 确保工作表名称始终包含 "mm2" 后缀
                sheet_name = f"{cross_section}mm2"
                
                if conductor_type == "铝导体":
                    write_data_to_excel(aluminum_workbook, sheet_name, [data_row], headers)
                elif conductor_type == "铜导体":
                    write_data_to_excel(copper_workbook, sheet_name, [data_row], headers)
                
            except Exception as e:
                errors.append(f"第 {index + 2} 行：处理失败，原因：{str(e)}，产品名称：{product_name}")
                # 将其他错误写入错误工作表
                data_row = [
                    row["产品编码"], row["工单"], row["生产日期"], row["订单日期"], 
                    row["订单"], row["单位名称"], product_name if not pd.isna(product_name) else "", row["型号"], 
                    row["数量"], row["分排"], row["交期"], row["绝缘"], 
                    row["成缆"], row["外护"], row["交期"], f"处理失败：{str(e)}"
                ]
                conductor_type = extract_conductor_type(str(row["产品名称"]) if not pd.isna(row["产品名称"]) else "")
                if conductor_type == "铝导体":
                    write_data_to_excel(aluminum_workbook, aluminum_error_sheet_name, [data_row], headers + ["错误信息"])
                elif conductor_type == "铜导体":
                   write_data_to_excel(copper_workbook, copper_error_sheet_name, [data_row], headers + ["错误信息"])
            
            processed_count += 1
            # 确保进度值在 0 到 1 之间
            progress = min((index + 1) / total_rows, 1.0)
            progress_bar.progress(progress)
            status_text.text(f'处理进度: {int(progress * 100)}%')

        aluminum_buffer = save_workbook_to_buffer(aluminum_workbook)
        copper_buffer = save_workbook_to_buffer(copper_workbook)

        return processed_count, errors, aluminum_buffer, copper_buffer
    except Exception as e:
        return 0, [f"文件处理失败，原因：{str(e)}"], None, None

def main():
    st.title("Excel订单数据分类工具")
    
    uploaded_file = st.file_uploader("上传你的 Excel 文件", type=["xls", "xlsx"])

    if uploaded_file is not None:
        if st.button("开始处理"):
            with st.spinner("正在处理数据..."):
                processed_count, errors, aluminum_buffer, copper_buffer = process_excel_data(uploaded_file)
            
            if errors:
                st.error("处理过程中出现以下错误：")
                for error in errors:
                    st.error(error)
            
            st.success(f"成功处理 {processed_count} 条记录！")
            st.success("数据分类完成！")
                
            col1, col2 = st.columns(2)
            with col1:
                if aluminum_buffer:
                    b64_al = base64.b64encode(aluminum_buffer.read()).decode()
                    href_al = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_al}" download="铝导体安排.xlsx">下载铝导体安排.xlsx</a>'
                    st.markdown(href_al, unsafe_allow_html=True)
            with col2:
                if copper_buffer:
                    b64_co = base64.b64encode(copper_buffer.read()).decode()
                    href_co = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_co}" download="铜导体安排.xlsx">下载铜导体安排.xlsx</a>'
                    st.markdown(href_co, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
